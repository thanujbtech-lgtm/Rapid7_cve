import re
import os

from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from playwright.sync_api import sync_playwright

# =========================================
# CONFIG
# =========================================

BASE_URL = "https://www.rapid7.com/db/"

FILE = "rapid7_cves.xlsx"

CVE_PATTERN = r"CVE-\d{4}-\d{4,7}"

# =========================================
# LOAD EXISTING ROWS
# =========================================

def load_existing_rows():

    if not os.path.exists(FILE):
        return []

    wb = load_workbook(FILE)

    ws = wb.active

    rows = []

    for row in ws.iter_rows(
        min_row=2,
        values_only=True
    ):

        if row[0]:
            rows.append(row)

    return rows

# =========================================
# DUPLICATE LOGIC
# SAME CVE + SAME DATE + SAME LINK
# =========================================

def load_existing_keys():

    rows = load_existing_rows()

    return {
        (
            str(r[0]),
            str(r[1]),
            str(r[4])
        )
        for r in rows
    }

# =========================================
# SAVE
# NO SORTING
# KEEP APPEND ORDER
# =========================================

def save_all(rows):

    wb = Workbook()

    ws = wb.active

    ws.append([
        "CVE",
        "DATE",
        "TITLE",
        "SEVERITY",
        "LINK"
    ])

    # =====================================
    # KEEP ORIGINAL ORDER
    # OLDEST EXISTING ROWS STAY TOP
    # NEWLY SCRAPED ROWS STAY BOTTOM
    # =====================================

    for r in rows:

        ws.append(r)

    wb.save(FILE)

# =========================================
# SCRAPE PAGE
# =========================================

def get_html(playwright):

    browser = playwright.chromium.launch(
        headless=True
    )

    page = browser.new_page()

    print("\nOpening Rapid7...\n")

    page.goto(
        BASE_URL,
        wait_until="domcontentloaded",
        timeout=120000
    )

    page.wait_for_timeout(5000)

    # =====================================
    # CLICK LOAD MORE
    # =====================================

    while True:

        try:

            load_more = page.locator(
                "text=Load more"
            )

            if load_more.count() > 0:

                print("Clicking Load More...")

                load_more.first.click()

                page.wait_for_timeout(3000)

            else:

                break

        except:

            break

    html = page.content()

    browser.close()

    return html

# =========================================
# MAIN
# =========================================

def main():

    existing_rows = load_existing_rows()

    existing_keys = load_existing_keys()

    new_rows = []

    with sync_playwright() as playwright:

        html = get_html(playwright)

    soup = BeautifulSoup(
        html,
        "html.parser"
    )

    cards = soup.find_all(
        ["a", "div", "article"]
    )

    print("\nProcessing cards...\n")

    for card in cards:

        text = card.get_text(
            " ",
            strip=True
        )

        cves = re.findall(
            CVE_PATTERN,
            text,
            re.IGNORECASE
        )

        if not cves:
            continue

        # =====================================
        # DATE
        # =====================================

        date_match = re.search(
            r"Published:\s*([0-9]{4}-[0-9]{2}-[0-9]{2})",
            text
        )

        if not date_match:
            continue

        publish_date = date_match.group(1)

        # =====================================
        # SEVERITY
        # =====================================

        severity_match = re.search(
            r"Severity:\s*([A-Za-z0-9]+)",
            text
        )

        severity = (
            severity_match.group(1)
            if severity_match
            else "UNKNOWN"
        )

        # =====================================
        # LINK
        # =====================================

        href = card.get("href")

        if href:

            if href.startswith("/"):

                link = (
                    "https://www.rapid7.com"
                    + href
                )

            else:

                link = href

        else:

            link = BASE_URL

        # =====================================
        # TITLE
        # =====================================

        title = text[:150]

        # =====================================
        # SAVE CVEs
        # =====================================

        for cve in cves:

            cve = cve.upper()

            key = (
                cve,
                publish_date,
                link
            )

            # =================================
            # SKIP DUPLICATES
            # =================================

            if key in existing_keys:
                continue

            print(
                "ADDING:",
                cve,
                publish_date
            )

            row = (
                cve,
                publish_date,
                title,
                severity,
                link
            )

            new_rows.append(row)

            existing_keys.add(key)

    # =========================================
    # APPEND NEW ROWS TO BOTTOM
    # =========================================

    all_rows = existing_rows + new_rows

    # =========================================
    # SAVE FILE
    # =========================================

    if new_rows:

        save_all(all_rows)

        print(
            f"\nAdded {len(new_rows)} new rows"
        )

        print(
            "\nNewest scraped rows appended at bottom"
        )

    else:

        print("\nNo new rows")

# =========================================
# RUN
# =========================================

if __name__ == "__main__":

    main()
